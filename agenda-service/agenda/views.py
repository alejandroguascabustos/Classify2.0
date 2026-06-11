"""
Equivalente Python de com.classify20.controller.AgendaController.

Endpoints (mismas rutas y mismo contrato que el controller Java):

  POST /guardar-agenda
      - Recibe los datos del formulario (mismos names del HTML: grado,
        grupo, profesor, materia, fecha, horaInicio, duracion, modalidad,
        temaPrincipal, objetivos, dificultades, materialesBasicos,
        recursosNecesarios).
      - Si llega con header X-Requested-With: XMLHttpRequest (el fetch de
        modalagenda.js) responde JSON: {"success": bool, "message": str}.
      - Si no, renderiza la plantilla agenda/agenda.html con "mensaje"
        o "error", igual que el flujo Thymeleaf.

  GET /programacion/exportarExcel
      - Genera programacion.xlsx con openpyxl replicando los estilos del
        Apache POI original (encabezado verde oscuro, fuente blanca en
        negrita, bordes finos, columnas autoajustadas).
"""
from datetime import date, datetime, time

from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET, require_POST
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import services
from .models import Agenda

AJAX_HEADER = "XMLHttpRequest"


# ── Helpers de parseo (equivalen al binding de @ModelAttribute) ──

def _str_o_none(valor):
    valor = (valor or "").strip()
    return valor or None


def _int_o_none(valor):
    try:
        return int(valor)
    except (TypeError, ValueError):
        return None


def _fecha_o_none(valor):
    try:
        return date.fromisoformat(valor)  # input type="date" → YYYY-MM-DD
    except (TypeError, ValueError):
        return None


def _hora_o_none(valor):
    if not valor:
        return None
    for formato in ("%H:%M:%S", "%H:%M"):  # input type="time" → HH:MM
        try:
            return datetime.strptime(valor, formato).time()
        except ValueError:
            continue
    return None


def _agenda_desde_formulario(post) -> Agenda:
    """Construye la entidad Agenda desde los datos del formulario."""
    return Agenda(
        grado=_int_o_none(post.get("grado")),
        grupo=_str_o_none(post.get("grupo")),
        profesor=_str_o_none(post.get("profesor")),
        materia=_str_o_none(post.get("materia")),
        fecha=_fecha_o_none(post.get("fecha")),
        hora_inicio=_hora_o_none(post.get("horaInicio")),
        duracion=_int_o_none(post.get("duracion")),
        modalidad=_str_o_none(post.get("modalidad")),
        tema_principal=_str_o_none(post.get("temaPrincipal")),
        objetivos=_str_o_none(post.get("objetivos")),
        dificultades=_str_o_none(post.get("dificultades")),
        materiales_basicos=_str_o_none(post.get("materialesBasicos")),
        recursos_necesarios=_str_o_none(post.get("recursosNecesarios")),
    )


# ── GET / (página de inicio del servicio, para ver en navegador) ─

@require_GET
def inicio(request):
    """Muestra el estado del servicio y las agendas registradas."""
    try:
        agendas = services.listar_agendas()
        return render(request, "agenda/inicio.html", {"agendas": agendas})
    except Exception as e:
        return render(request, "agenda/inicio.html", {"error_bd": str(e)})


# ── POST /guardar-agenda ─────────────────────────────────────────

@csrf_exempt  # el formulario original no envía token CSRF de Django
@require_POST
def guardar_agenda(request):
    es_ajax = request.headers.get("X-Requested-With") == AJAX_HEADER

    try:
        agenda = _agenda_desde_formulario(request.POST)
        services.guardar_agenda(agenda)

        if es_ajax:
            return JsonResponse({"success": True, "message": "¡Agenda guardada!"})

        return render(
            request,
            "agenda/agenda.html",
            {"mensaje": "¡Agenda guardada exitosamente!"},
        )

    except Exception as e:  # mismo manejo genérico del controller Java
        if es_ajax:
            return JsonResponse(
                {"success": False, "message": f"Error al guardar: {e}"},
                status=500,
            )
        return render(
            request,
            "agenda/agenda.html",
            {"error": "Error al guardar la agenda."},
        )


# ── GET /programacion/exportarExcel ──────────────────────────────

@require_GET
def exportar_excel(request):
    agendas = services.listar_agendas()

    wb = Workbook()
    ws = wb.active
    ws.title = "Programación"

    # Estilos equivalentes a los de Apache POI
    borde_fino = Border(
        bottom=Side(style="thin"),
        top=Side(style="thin"),
        left=Side(style="thin"),
        right=Side(style="thin"),
    )
    # IndexedColors.DARK_GREEN de POI ≈ 006100
    relleno_encabezado = PatternFill(
        start_color="006100", end_color="006100", fill_type="solid"
    )
    fuente_encabezado = Font(color="FFFFFF", bold=True)
    centrado = Alignment(horizontal="center")

    encabezados = [
        "ID", "Materia", "Profesor", "Fecha", "Hora Inicio",
        "Curso", "Modalidad", "Tema", "Duracion",
    ]
    for col, titulo in enumerate(encabezados, start=1):
        celda = ws.cell(row=1, column=col, value=titulo)
        celda.fill = relleno_encabezado
        celda.font = fuente_encabezado
        celda.alignment = centrado
        celda.border = borde_fino

    for fila, a in enumerate(agendas, start=2):
        curso = f"{a.grado if a.grado is not None else ''} {a.grupo or ''}".strip()
        valores = [
            a.id if a.id is not None else 0,
            a.materia or "",
            a.profesor or "",
            a.fecha.isoformat() if a.fecha else "",
            a.hora_inicio.isoformat() if a.hora_inicio else "",
            curso,
            a.modalidad or "",
            a.tema_principal or "",
            f"{a.duracion} min" if a.duracion is not None else "",
        ]
        for col, valor in enumerate(valores, start=1):
            celda = ws.cell(row=fila, column=col, value=valor)
            celda.border = borde_fino

    # Autoajustar ancho de columnas (equivale a sheet.autoSizeColumn)
    for col in range(1, len(encabezados) + 1):
        letra = get_column_letter(col)
        ancho = max(
            len(str(celda.value)) if celda.value is not None else 0
            for celda in ws[letra]
        )
        ws.column_dimensions[letra].width = ancho + 3

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=programacion.xlsx"
    wb.save(response)
    return response
